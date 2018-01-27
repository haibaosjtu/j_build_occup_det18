import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
from sklearn.svm import SVR
import numpy as np
import matplotlib.pyplot as plt

wb = openpyxl.load_workbook('Data.xlsx')
sheet = wb.get_sheet_by_name('Sheet3')


#for i in range(1,151,1)     
#    for j in range(1,5,1)
#        temp=temp+[sheet.cell(row=i,column=j)]
#    
data = [[0 for x in range(3)] for x in range(2880)] 
for i in range(2,2882,1):   
    for j in range(2,5,1):
        data[i-2][j-2]=sheet.cell(row=i,column=j).value

label=[0 for x in range(2880)]  
for j in range(2,2882,1):
    label[j-2]=sheet.cell(row=j,column=5).value

train=data[0:2160]
trainLabel=label[0:2160]

test=data[2160:2880]
testLabel=label[2160:2880]
    

