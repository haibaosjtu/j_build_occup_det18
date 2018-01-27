import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('Fish.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')


#for i in range(1,151,1)     
#    for j in range(1,5,1)
#        temp=temp+[sheet.cell(row=i,column=j)]
#    
data = [[0 for x in range(4)] for x in range(150)] 
for i in range(1,151,1):   
    for j in range(1,5,1):
        data[i-1][j-1]=sheet.cell(row=i,column=j).value
    
    

print data
  