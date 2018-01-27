from __future__ import division
import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
from sklearn.svm import SVR
import numpy as np
import matplotlib.pyplot as plt

wb = openpyxl.load_workbook('SmartBuilding3.xlsx')
sheet = wb.get_sheet_by_name('Sheet2')


#for i in range(1,151,1)     
#    for j in range(1,5,1)
#        temp=temp+[sheet.cell(row=i,column=j)]
#    
data = [[0 for x in range(4)] for x in range(2880)] 
for i in range(2,2882,1):   
    for j in range(2,6,1):
        data[i-2][j-2]=sheet.cell(row=i,column=j).value

label=[0 for x in range(2880)]  
for j in range(2,2882,1):
    label[j-2]=sheet.cell(row=j,column=7).value

train=data[0:2160]
trainLabel=label[0:2160]

test=data[2160:2880]
testLabel=label[2160:2880]
    


        
X=train
y=trainLabel
C=100
epsilon=0.01
gamma=0.001
print 'C= '+str(C)+' epsilon= '+str(epsilon)+' gamma= '+str(gamma)
print 'train process for season\n'

#clf = SVR(C=1000, epsilon=0.01)
#temp1=clf.fit(X, y).predict(train)
clf2 = SVR(kernel='rbf', C=C, gamma=gamma,epsilon=epsilon)
temp2=clf2.fit(X, y).predict(train)
#clf3 = SVR(kernel='poly', C=1e3, degree=2)
#temp3=clf3.fit(X, y).predict(test)



#clf2=SVR(C=1.0, cache_size=200, coef0=0.0, degree=3, epsilon=0.2, gamma='auto',
#    kernel='rbf', max_iter=-1, shrinking=True, tol=0.001, verbose=False)
#temp2=clf2.fit(X,y).predict(X)
#plt.hold('on')
plt.plot(trainLabel, c='y', label='Data')
#plt.plot(temp1, c='b', label='Linear model')
plt.plot(temp2, c='g', label='RBF model')
#plt.plot(temp3, c='m', label='Poly model')
#plt.plot(temp4, c='r', label='Prediction')
plt.xlabel('data')
plt.ylabel('target')
#plt.title('Support Vector Regression')
plt.legend()
plt.show()


comparison=abs(temp2-trainLabel)
aveError=sum(comparison)/len(comparison)
maxError=max(comparison)
print 'aveError '+str(aveError)
print 'maxError '+str(maxError)

ratio=aveError/12
print 'ErrRatio  '+str(ratio)

num=0;
for i in range(0,len(comparison),1):
    if comparison[i]>0.5:
        num=num+1

errorRate=float(num/len(comparison))
print 'errorRate '+str(errorRate)