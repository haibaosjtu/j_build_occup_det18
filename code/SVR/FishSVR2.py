import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
from sklearn.svm import SVR
import numpy as np
import matplotlib.pyplot as plt

wb = openpyxl.load_workbook('Fish.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')


#for i in range(1,151,1)     
#    for j in range(1,5,1)
#        temp=temp+[sheet.cell(row=i,column=j)]
#    
data = [[0 for x in range(4)] for x in range(150)] 
for i in range(1,151,1):   
    for j in range(1,5,1):
        data[i-1][j-1]=sheet.cell(row=i,column=j).value

group=[0 for x in range(150)]
for i in range(0,150):
    if i<=49:
        group[i]=1
    elif i>49 and i<=99:
        group[i]=2
    else:
        group[i]=3
 

train=data[10:50]+data[60:100]+data[110:150]
trainLabel=group[10:50]+group[60:100]+group[110:150]

test=data[0:10]+data[50:60]+data[100:110]
testLabel=group[0:10]+group[50:60]+group[100:110]


       
X=train
y=trainLabel

clf = SVR(C=1.0, epsilon=0.2)
temp1=clf.fit(X, y).predict(test)
clf2 = SVR(kernel='rbf', C=1e3, gamma=0.1)
temp2=clf2.fit(X, y).predict(test)
clf3 = SVR(kernel='poly', C=1e3, degree=2)
temp3=clf3.fit(X, y).predict(test)

  

#clf2=SVR(C=1.0, cache_size=200, coef0=0.0, degree=3, epsilon=0.2, gamma='auto',
#    kernel='rbf', max_iter=-1, shrinking=True, tol=0.001, verbose=False)
#temp2=clf2.fit(X,y).predict(X)
plt.hold('on')
plt.plot(testLabel, c='y', label='Data')
plt.plot(temp1, c='b', label='Linear model')
plt.plot(temp2, c='g', label='RBF model')
plt.plot(temp3, c='m', label='Poly model')

plt.xlabel('data')
plt.ylabel('target')
#plt.title('Support Vector Regression')
plt.legend()
plt.show()
  